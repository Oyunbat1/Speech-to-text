from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.diarize import Segment

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="FFDDDDDD", end_color="FFDDDDDD", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFFFF59D", end_color="FFFFF59D", fill_type="solid")
WARNING_FONT = Font(bold=True, color="FF8B6E00")

HEADERS = ["Start", "End", "Role", "Text"]
WARNING_MESSAGE = (
    "Teacher voice not confidently matched - roles assigned by speaking time."
)


def _format_time(seconds: float) -> str:
    total = int(round(seconds))
    h, rem = divmod(total, 3600)
    m, s = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def write_xlsx(
    segments: list[Segment],
    texts: list[str],
    out_path: str,
    fallback_warning: bool = False,
) -> None:
    """Columns: Start (HH:MM:SS) | End (HH:MM:SS) | Role | Text."""
    if len(segments) != len(texts):
        raise ValueError(
            f"segments ({len(segments)}) and texts ({len(texts)}) length mismatch"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Transcript"

    row_offset = 1
    if fallback_warning:
        ws.cell(row=1, column=1, value=WARNING_MESSAGE)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
        warn_cell = ws.cell(row=1, column=1)
        warn_cell.fill = WARNING_FILL
        warn_cell.font = WARNING_FONT
        warn_cell.alignment = Alignment(horizontal="left", vertical="center")
        row_offset = 2

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=row_offset, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for i, (seg, text) in enumerate(zip(segments, texts)):
        row = row_offset + 1 + i
        ws.cell(row=row, column=1, value=_format_time(seg.start))
        ws.cell(row=row, column=2, value=_format_time(seg.end))
        ws.cell(row=row, column=3, value=seg.role or "")
        text_cell = ws.cell(row=row, column=4, value=text)
        text_cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {1: 12, 2: 12, 3: 10, 4: 80}
    for col_idx, header in enumerate(HEADERS, start=1):
        max_len = len(header)
        for row_cells in ws.iter_rows(
            min_row=row_offset, max_row=ws.max_row, min_col=col_idx, max_col=col_idx
        ):
            for cell in row_cells:
                if cell.value is None:
                    continue
                length = len(str(cell.value))
                if length > max_len:
                    max_len = length
        ws.column_dimensions[ws.cell(row=row_offset, column=col_idx).column_letter].width = min(
            max(widths.get(col_idx, 12), max_len + 2), widths.get(col_idx, 12) if col_idx < 4 else 100
        )

    wb.save(out_path)
