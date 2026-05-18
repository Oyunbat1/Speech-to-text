"""End-to-end smoke test for the lecture-transcribe pipeline.

Requires:
  - ffmpeg on PATH
  - HF_TOKEN and CHIMEGE_TOKEN set in environment / .env
  - tests/fixtures/sample.mp4 - ~30s clip with two clearly distinct voices
  - tests/fixtures/teacher_ref.wav - ~10s of just the teacher's voice
"""

from pathlib import Path

import pytest
from dotenv import load_dotenv
from openpyxl import load_workbook

from app.pipeline import process_lecture

load_dotenv()

FIXTURES = Path(__file__).parent / "fixtures"
SAMPLE_LECTURE = FIXTURES / "sample.mp4"
TEACHER_REF = FIXTURES / "teacher_ref.wav"


@pytest.mark.skipif(
    not SAMPLE_LECTURE.exists() or not TEACHER_REF.exists(),
    reason="test fixtures missing - add sample.mp4 and teacher_ref.wav under tests/fixtures/",
)
def test_pipeline_end_to_end(tmp_path: Path) -> None:
    output = tmp_path / "out.xlsx"

    process_lecture(
        lecture_path=str(SAMPLE_LECTURE),
        teacher_reference_path=str(TEACHER_REF),
        output_xlsx_path=str(output),
    )

    assert output.exists(), "xlsx output missing"

    wb = load_workbook(output)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) > 1, "expected header + data rows"

    data_rows = [
        r for r in rows[1:]
        if r[2] == "Teacher" or (isinstance(r[2], str) and r[2].startswith("Student"))
    ]
    assert len(data_rows) > 0, "no data rows with role"

    roles = {r[2] for r in data_rows}
    assert "Teacher" in roles, "no Teacher row"
    assert any(r.startswith("Student") for r in roles), "no Student row"
